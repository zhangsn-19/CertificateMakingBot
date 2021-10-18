import openpyxl
from openpyxl.styles import Font, colors, PatternFill  
# 导入样式的字体大小和颜色函数（注意大小写）
# 背景色设置函数（注意大小写）
def deletecolumns(sheet,column_num):
  for column in range(column_num, sheet.max_column - 1):
    for row in range(1,sheet.max_row):
      sheet[row][column].value = sheet[row][column + 1].value
  for cell in list(sheet.columns)[sheet.max_column - 1]:
    cell.value = None


table = openpyxl.load_workbook('小伙伴早安第一期.xlsx')
sheet = table['打卡活动'] #选定名为打开活动的表单进行操作

deletecolumns(sheet, 1)
deletecolumns(sheet, 2)
deletecolumns(sheet, 2)
deletecolumns(sheet, 3)

sheet['A1'].font = Font(#color = colors.RED, #使用预先置好的颜色变量
                        name = u"微软雅黑",
						color ="82318E", #设置字体颜色
						size = 12, #设置文字大小
						bold = True, #设定粗体
						italic = False #设置斜体
						)
sheet['B1'].font = Font(
                        name = "Arial",
                        color = "82318E",
                        size = 12, 
                        bold = False
                        )
sheet['C1'].font = Font(name = "Arial",
                        color = "82318E",
                        size = 12, 
                        bold = False) #使用RGB数字表示颜色

#指定整列风格，(openpyxl有整行替换的函数，但使用不出来，所以这里使用递归法方式一个一个换)
fill = PatternFill("solid", fgColor="31B404")
for y in range(1,4):
    for x in range(1,sheet.max_row//2):
	    sheet.cell(row=x*2,column=y).fill = fill #将第二列的样式进行逐一替换
sheet.title="第一期打卡活动统计表"
table.save('小伙伴早安第一期.xlsx')#最后一段要保存文件
for x in range(1,sheet.max_row):
    if sheet.cell(row=3,column=x).value != "21":
        sheet.cell(row=1,column=x).value = None
        sheet.cell(row=2,column=x).value = None